import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Color palette ──────────────────────────────────────────────────────────────
NAVY   = "052D5C"
BLUE   = "2D71AE"
RED    = "D91D23"
WHITE  = "FFFFFF"
LIGHT  = "EEF4FB"
GRAY   = "F5F7FA"
YELLOW = "FFF9C4"
GREEN  = "E8F5E9"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="1A1A2E"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — INTERIOR PAINTING
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Interior Painting"

# Column widths
col_widths = [28, 18, 18, 18, 22, 40]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title row
ws1.row_dimensions[1].height = 36
ws1.merge_cells("A1:F1")
ws1["A1"] = "SGB Custom Painting — Interior Painting Pricing Variables"
ws1["A1"].fill = fill(NAVY)
ws1["A1"].font = hdr_font(14)
ws1["A1"].alignment = center()

# Subtitle
ws1.row_dimensions[2].height = 22
ws1.merge_cells("A2:F2")
ws1["A2"] = "Current prices are Northern California / Chico area averages. Scott: please review and update the 'Scott's Price' column with your actual numbers."
ws1["A2"].fill = fill(BLUE)
ws1["A2"].font = Font(name="Calibri", size=10, color=WHITE, italic=True)
ws1["A2"].alignment = center()

# Section: Room Pricing
ws1.row_dimensions[3].height = 8

ws1.row_dimensions[4].height = 28
ws1.merge_cells("A4:F4")
ws1["A4"] = "SECTION 1 — PER ROOM PRICES  (includes walls, ceiling, trim & doors in that room)"
ws1["A4"].fill = fill(BLUE)
ws1["A4"].font = hdr_font(11)
ws1["A4"].alignment = left()

# Column headers
ws1.row_dimensions[5].height = 30
headers = ["Room Type", "Good Tier ($)", "Better Tier ($)", "Best Tier ($)", "Scott's Price (Better)", "Notes"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=5, column=col, value=h)
    c.fill = fill(NAVY)
    c.font = hdr_font(10)
    c.alignment = center()
    c.border = border()

# Room data: [label, good, better, best, note]
rooms = [
    ("Small Bedroom (< 120 sq ft)",     450, 550, 700,  "Walls + ceiling + trim + 1 door"),
    ("Standard Bedroom (120–180 sq ft)", 550, 675, 850,  "Walls + ceiling + trim + 1 door"),
    ("Master Bedroom (180–300 sq ft)",   700, 875, 1100, "Walls + ceiling + trim + 2 doors"),
    ("Living Room",                      750, 925, 1150, "Open concept; may vary by size"),
    ("Dining Room",                      550, 675, 850,  "Often combined with living room"),
    ("Kitchen",                          600, 750, 950,  "Excludes cabinets; walls + ceiling"),
    ("Small Bathroom",                   350, 425, 550,  "Walls + ceiling + trim"),
    ("Master Bathroom",                  500, 625, 775,  "Larger space; more detail work"),
    ("Hallway / Entry",                  350, 425, 550,  "Per hallway section"),
    ("Office / Den",                     550, 675, 850,  "Similar to standard bedroom"),
    ("Stairwell / Landing",              450, 550, 700,  "Includes stair walls + ceiling"),
]

for i, (label, good, better, best, note) in enumerate(rooms, 6):
    row_fill = fill(LIGHT) if i % 2 == 0 else fill(WHITE)
    ws1.row_dimensions[i].height = 22
    data = [label, good, better, best, "", note]
    for col, val in enumerate(data, 1):
        c = ws1.cell(row=i, column=col, value=val)
        c.fill = row_fill
        c.font = body_font(10, bold=(col == 1))
        c.alignment = center() if col in [2,3,4,5] else left()
        c.border = border()
        if col == 5:
            c.fill = fill(YELLOW)
            c.font = body_font(10, bold=True, color="333300")

# Section: Prep Add-Ons
start = len(rooms) + 7
ws1.row_dimensions[start].height = 8
ws1.row_dimensions[start+1].height = 28
ws1.merge_cells(f"A{start+1}:F{start+1}")
ws1[f"A{start+1}"] = "SECTION 2 — PREP WORK ADD-ONS  (added to room total when selected)"
ws1[f"A{start+1}"].fill = fill(BLUE)
ws1[f"A{start+1}"].font = hdr_font(11)
ws1[f"A{start+1}"].alignment = left()

# Add-on headers
hrow = start + 2
ws1.row_dimensions[hrow].height = 30
addons_headers = ["Add-On Item", "Flat Price ($)", "Scott's Price", "Notes"]
for col, h in enumerate(addons_headers, 1):
    c = ws1.cell(row=hrow, column=col, value=h)
    c.fill = fill(NAVY)
    c.font = hdr_font(10)
    c.alignment = center()
    c.border = border()

addons = [
    ("Wallpaper Removal (per room)",          450, "Includes scoring, stripping, wall prep"),
    ("Popcorn Ceiling Removal (per room)",     400, "Scrape, skim coat, sand smooth"),
    ("Minor Drywall Repairs (per room)",       150, "Small holes, nail pops, hairline cracks"),
    ("Moderate Drywall Repairs (per room)",    350, "Larger patches, water damage areas"),
]

for i, (label, flat, note) in enumerate(addons, hrow+1):
    row_fill = fill(LIGHT) if i % 2 == 0 else fill(WHITE)
    ws1.row_dimensions[i].height = 22
    data = [label, flat, "", note]
    for col, val in enumerate(data, 1):
        c = ws1.cell(row=i, column=col, value=val)
        c.fill = row_fill
        c.font = body_font(10, bold=(col == 1))
        c.alignment = center() if col in [2,3] else left()
        c.border = border()
        if col == 3:
            c.fill = fill(YELLOW)
            c.font = body_font(10, bold=True, color="333300")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — EXTERIOR PAINTING
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Exterior Painting")

col_widths2 = [30, 18, 18, 18, 22, 40]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.row_dimensions[1].height = 36
ws2.merge_cells("A1:F1")
ws2["A1"] = "SGB Custom Painting — Exterior Painting Pricing Variables"
ws2["A1"].fill = fill(NAVY)
ws2["A1"].font = hdr_font(14)
ws2["A1"].alignment = center()

ws2.row_dimensions[2].height = 22
ws2.merge_cells("A2:F2")
ws2["A2"] = "Prices are based on home size tiers. Scott: update 'Scott's Price' column with your actual numbers."
ws2["A2"].fill = fill(BLUE)
ws2["A2"].font = Font(name="Calibri", size=10, color=WHITE, italic=True)
ws2["A2"].alignment = center()

ws2.row_dimensions[3].height = 8
ws2.row_dimensions[4].height = 28
ws2.merge_cells("A4:F4")
ws2["A4"] = "SECTION 1 — BASE PRICE BY HOME SIZE  (full exterior: siding, trim, doors, windows)"
ws2["A4"].fill = fill(BLUE)
ws2["A4"].font = hdr_font(11)
ws2["A4"].alignment = left()

ws2.row_dimensions[5].height = 30
for col, h in enumerate(["Home Size", "Good Tier ($)", "Better Tier ($)", "Best Tier ($)", "Scott's Price (Better)", "Notes"], 1):
    c = ws2.cell(row=5, column=col, value=h)
    c.fill = fill(NAVY)
    c.font = hdr_font(10)
    c.alignment = center()
    c.border = border()

ext_sizes = [
    ("Small  (under 1,200 sq ft)",        1800, 2400, 3200, "1-story, simple profile"),
    ("Medium (1,200 – 2,000 sq ft)",       2800, 3600, 4800, "Typical single-family home"),
    ("Large  (2,000 – 2,800 sq ft)",       3800, 4900, 6500, "Two-story or larger footprint"),
    ("Extra Large (2,800+ sq ft)",         5200, 6800, 9000, "Custom quote recommended"),
]

for i, (label, good, better, best, note) in enumerate(ext_sizes, 6):
    row_fill = fill(LIGHT) if i % 2 == 0 else fill(WHITE)
    ws2.row_dimensions[i].height = 22
    for col, val in enumerate([label, good, better, best, "", note], 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.fill = row_fill
        c.font = body_font(10, bold=(col == 1))
        c.alignment = center() if col in [2,3,4,5] else left()
        c.border = border()
        if col == 5:
            c.fill = fill(YELLOW)
            c.font = body_font(10, bold=True, color="333300")

# Exterior add-ons
estart = 11
ws2.row_dimensions[estart].height = 8
ws2.row_dimensions[estart+1].height = 28
ws2.merge_cells(f"A{estart+1}:F{estart+1}")
ws2[f"A{estart+1}"] = "SECTION 2 — EXTERIOR PREP ADD-ONS"
ws2[f"A{estart+1}"].fill = fill(BLUE)
ws2[f"A{estart+1}"].font = hdr_font(11)
ws2[f"A{estart+1}"].alignment = left()

ehrow = estart + 2
ws2.row_dimensions[ehrow].height = 30
for col, h in enumerate(["Add-On Item", "Flat Price ($)", "Scott's Price", "Notes"], 1):
    c = ws2.cell(row=ehrow, column=col, value=h)
    c.fill = fill(NAVY)
    c.font = hdr_font(10)
    c.alignment = center()
    c.border = border()

ext_addons = [
    ("Power Washing (whole house)",           275,  "Included by default; can be removed"),
    ("Wood Rot Repair (minor, per area)",      350,  "Small sections of siding/trim"),
    ("Wood Rot Repair (moderate, per area)",   700,  "Larger board replacement"),
    ("Caulking & Sealing (full house)",        200,  "Windows, doors, trim gaps"),
    ("Deck / Patio Painting or Staining",      1000, "Per standard deck up to 400 sq ft"),
    ("Stair Railing / Patio Railing",          275,  "Per railing section"),
    ("Fireplace Brick Painting",               400,  "Exterior brick fireplace"),
]

for i, (label, flat, note) in enumerate(ext_addons, ehrow+1):
    row_fill = fill(LIGHT) if i % 2 == 0 else fill(WHITE)
    ws2.row_dimensions[i].height = 22
    for col, val in enumerate([label, flat, "", note], 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.fill = row_fill
        c.font = body_font(10, bold=(col == 1))
        c.alignment = center() if col in [2,3] else left()
        c.border = border()
        if col == 3:
            c.fill = fill(YELLOW)
            c.font = body_font(10, bold=True, color="333300")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — CABINET PAINTING
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Cabinet Painting")

col_widths3 = [30, 18, 18, 18, 22, 40]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.row_dimensions[1].height = 36
ws3.merge_cells("A1:F1")
ws3["A1"] = "SGB Custom Painting — Cabinet Painting Pricing Variables"
ws3["A1"].fill = fill(NAVY)
ws3["A1"].font = hdr_font(14)
ws3["A1"].alignment = center()

ws3.row_dimensions[2].height = 22
ws3.merge_cells("A2:F2")
ws3["A2"] = "Prices are per door or per drawer. Scott: update 'Scott's Price' column with your actual numbers."
ws3["A2"].fill = fill(BLUE)
ws3["A2"].font = Font(name="Calibri", size=10, color=WHITE, italic=True)
ws3["A2"].alignment = center()

ws3.row_dimensions[3].height = 8
ws3.row_dimensions[4].height = 28
ws3.merge_cells("A4:F4")
ws3["A4"] = "SECTION 1 — PER UNIT PRICES  (includes prep, prime, and 2 finish coats)"
ws3["A4"].fill = fill(BLUE)
ws3["A4"].font = hdr_font(11)
ws3["A4"].alignment = left()

ws3.row_dimensions[5].height = 30
for col, h in enumerate(["Item", "Good Tier ($)", "Better Tier ($)", "Best Tier ($)", "Scott's Price (Better)", "Notes"], 1):
    c = ws3.cell(row=5, column=col, value=h)
    c.fill = fill(NAVY)
    c.font = hdr_font(10)
    c.alignment = center()
    c.border = border()

cab_items = [
    ("Cabinet Door (per door)",              85,  110, 145, "Both sides; includes prep + prime"),
    ("Drawer Front (per drawer front)",       55,   70,  90, "Front face only"),
    ("Cabinet Box / Frame (per linear ft)",   35,   45,  60, "Interior and exterior of box"),
]

for i, (label, good, better, best, note) in enumerate(cab_items, 6):
    row_fill = fill(LIGHT) if i % 2 == 0 else fill(WHITE)
    ws3.row_dimensions[i].height = 22
    for col, val in enumerate([label, good, better, best, "", note], 1):
        c = ws3.cell(row=i, column=col, value=val)
        c.fill = row_fill
        c.font = body_font(10, bold=(col == 1))
        c.alignment = center() if col in [2,3,4,5] else left()
        c.border = border()
        if col == 5:
            c.fill = fill(YELLOW)
            c.font = body_font(10, bold=True, color="333300")

# Cabinet add-ons
cstart = 10
ws3.row_dimensions[cstart].height = 8
ws3.row_dimensions[cstart+1].height = 28
ws3.merge_cells(f"A{cstart+1}:F{cstart+1}")
ws3[f"A{cstart+1}"] = "SECTION 2 — CABINET ADD-ONS"
ws3[f"A{cstart+1}"].fill = fill(BLUE)
ws3[f"A{cstart+1}"].font = hdr_font(11)
ws3[f"A{cstart+1}"].alignment = left()

chrow = cstart + 2
ws3.row_dimensions[chrow].height = 30
for col, h in enumerate(["Add-On Item", "Flat Price ($)", "Scott's Price", "Notes"], 1):
    c = ws3.cell(row=chrow, column=col, value=h)
    c.fill = fill(NAVY)
    c.font = hdr_font(10)
    c.alignment = center()
    c.border = border()

cab_addons = [
    ("Hardware Removal & Reinstall (flat fee)",  100, "Remove hinges/pulls, reinstall after"),
    ("Glass Door Masking (per door)",              20, "Tape off glass panels"),
    ("Glaze / Specialty Finish (per door)",        50, "Antiquing, glaze, or two-tone finish"),
]

for i, (label, flat, note) in enumerate(cab_addons, chrow+1):
    row_fill = fill(LIGHT) if i % 2 == 0 else fill(WHITE)
    ws3.row_dimensions[i].height = 22
    for col, val in enumerate([label, flat, "", note], 1):
        c = ws3.cell(row=i, column=col, value=val)
        c.fill = row_fill
        c.font = body_font(10, bold=(col == 1))
        c.alignment = center() if col in [2,3] else left()
        c.border = border()
        if col == 3:
            c.fill = fill(YELLOW)
            c.font = body_font(10, bold=True, color="333300")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — INSTRUCTIONS FOR SCOTT
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Instructions for Scott")

ws4.column_dimensions["A"].width = 80
ws4.row_dimensions[1].height = 40
ws4["A1"] = "Instructions for Scott — How to Use This Spreadsheet"
ws4["A1"].fill = fill(NAVY)
ws4["A1"].font = hdr_font(16)
ws4["A1"].alignment = left()

instructions = [
    ("", ""),
    ("PURPOSE", "This spreadsheet contains all the pricing variables used in your online Painting Cost Estimator on your website."),
    ("", "When a potential customer uses the calculator, these are the exact numbers that drive the estimate they see."),
    ("", ""),
    ("HOW TO REVIEW", "1. Go through each tab: Interior Painting, Exterior Painting, Cabinet Painting"),
    ("", "2. Look at the 'Better Tier' column — this is the most commonly selected tier and the most important one to get right."),
    ("", "3. The yellow 'Scott's Price (Better)' column is where YOU fill in your actual numbers."),
    ("", "4. If you are happy with the Good and Best tier numbers too, fill those in as well."),
    ("", "5. Leave any cell blank if you want to keep the current Northern CA average."),
    ("", ""),
    ("TIERS EXPLAINED", "Good = Entry-level paint (Sherwin-Williams ProMar 200 range). Lowest price, solid quality."),
    ("", "Better = Mid-grade paint (Sherwin-Williams SuperPaint / Duration range). Most popular choice."),
    ("", "Best = Premium paint (Sherwin-Williams Emerald / Benjamin Moore Aura range). Top of the line."),
    ("", ""),
    ("HOW PRICES ARE USED", "Interior: Customer selects rooms. Each room's price is added up. Add-ons are added on top."),
    ("", "Exterior: Customer selects home size. One base price is used. Add-ons are added on top."),
    ("", "Cabinet: Customer enters number of doors and drawers. Price per door x doors + price per drawer x drawers."),
    ("", ""),
    ("AFTER YOU FILL IT IN", "Send the completed spreadsheet back to Peterson SEO."),
    ("", "We will update the calculator on your website within 24 hours."),
    ("", ""),
    ("QUESTIONS?", "Contact Peterson SEO at www.petersonseoconsulting.com"),
]

for i, (label, text) in enumerate(instructions, 2):
    ws4.row_dimensions[i].height = 20
    if label:
        ws4.cell(row=i, column=1, value=f"  {label}:   {text}").font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    else:
        ws4.cell(row=i, column=1, value=f"              {text}").font = Font(name="Calibri", size=10, color="1A1A2E")
    ws4.cell(row=i, column=1).alignment = left()

# Save
out_path = "/home/ubuntu/sgb_calculator/SGB_Painting_Pricing_Variables.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
